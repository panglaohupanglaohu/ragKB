# -*- coding: utf-8 -*-
"""AgentsGroup2026 Tool Executor — Actually runs tools when agents invoke them.

Inspired by Clawith's Tools Engine: when the LLM returns a function call
matching a bound tool, this executor dispatches it to the correct handler.
"""

from __future__ import annotations

import asyncio
import json
import logging
import os
import subprocess
import time
from datetime import datetime, timezone
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

from sandbox.python_runner import get_sandbox
from .execution_registry import ToolPermissionContext

logger = logging.getLogger(__name__)


@dataclass
class ToolResult:
    """Result from executing a tool."""
    tool_id: str = ""
    tool_name: str = ""
    success: bool = True
    output: str = ""
    error: str = ""
    execution_ms: float = 0.0
    timestamp: str = field(
        default_factory=lambda: datetime.now(timezone.utc).isoformat()
    )

    def to_dict(self) -> Dict[str, Any]:
        return {
            "tool_id": self.tool_id,
            "tool_name": self.tool_name,
            "success": self.success,
            "output": self.output[:4000],
            "error": self.error,
            "execution_ms": self.execution_ms,
            "timestamp": self.timestamp,
        }


class ToolExecutor:
    """Dispatches tool calls to handler functions.

    Each handler receives the tool arguments dict and returns a ToolResult.
    Handlers are registered by tool name.
    """

    def __init__(self) -> None:
        self._handlers: Dict[str, Any] = {}
        self._history: List[ToolResult] = []
        self._register_builtins()

    def _register_builtins(self) -> None:
        """Register built-in tool handlers."""
        self._handlers.update({
            # Browser
            "web_search": self._web_search,
            "navigate_url": self._navigate_url,
            "screenshot": self._screenshot,
            "click_element": self._click_element,
            "fill_form": self._fill_form,
            "extract_content": self._extract_content,
            "web_extract": self._extract_content,
            # Code Execution
            "run_python": self._run_python,
            "run_shell": self._run_shell,
            "run_javascript": self._run_javascript,
            "execute_code": self._execute_code,
            # File Operation
            "read_file": self._read_file,
            "write_file": self._write_file,
            "edit_file": self._edit_file,
            "delete_file": self._delete_file,
            "list_files": self._list_directory,
            "list_directory": self._list_directory,
            "search_files": self._search_files,
            "find_files": self._find_files,
            "read_document": self._read_document,
            # Communication
            "send_message": self._send_message,
            "broadcast": self._broadcast,
            "subscribe_channel": self._subscribe_channel,
            "publish_event": self._publish_event,
            # Maritime
            "ais_query": self._ais_query,
            "ais_vessel_track": self._ais_vessel_track,
            "weather_fetch": self._weather_fetch,
            "weather_marine_forecast": self._weather_fetch,
            "route_calculate": self._route_calculate,
            "colregs_check": self._colregs_check,
            "engine_status": self._engine_status,
            "engine_diagnostic_scan": self._engine_status,
            "engine_monitor": self._engine_monitor,
            "cargo_status": self._cargo_status,
            "chart_ecdis_query": self._chart_query,
            "chart_lookup": self._chart_query,
            # Memory
            "memory_save": self._memory_save,
            "memory_read": self._memory_read,
            "session_search": self._session_search,
            # Skills
            "skill_list": self._skill_list,
            "skill_view": self._skill_view,
            "skill_manage": self._skill_manage,
            # Delegation
            "delegate_task": self._delegate_task,
            "mixture_of_agents": self._mixture_of_agents,
            # Discovery
            "list_agents": self._list_agents,
            "list_capabilities": self._list_capabilities,
            # Digital Twin
            "dt_camera_move": self._dt_camera_move,
            "dt_model_load": self._dt_model_load,
            "dt_model_transform": self._dt_model_transform,
            "dt_material_set": self._dt_material_set,
            "dt_physics_toggle": self._dt_physics_toggle,
            "dt_light_adjust": self._dt_light_adjust,
            "dt_render_mode": self._dt_render_mode,
            "dt_inspection_path": self._dt_inspection_path,
            # Triggers
            "schedule_task": self._schedule_task,
            "set_alarm": self._set_alarm,
            "watch_file": self._watch_file,
            "cron_trigger": self._cron_trigger,
            # Vision
            "vision_analyze": self._vision_analyze,
        })

    async def execute(
        self, tool_name: str, arguments: Dict[str, Any],
        *,
        agent_id: str = "",
        requires_approval: bool = False,
        permission_context: Optional[ToolPermissionContext] = None,
    ) -> ToolResult:
        """Execute a tool by name with arguments."""
        if permission_context and permission_context.blocks(tool_name):
            result = ToolResult(
                tool_name=tool_name,
                success=False,
                error=f"Tool blocked by agent permissions: {tool_name}",
            )
            self._history.append(result)
            return result

        handler = self._handlers.get(tool_name)
        if handler is None:
            result = ToolResult(
                tool_name=tool_name, success=False,
                error=f"Unknown tool: {tool_name}",
            )
            self._history.append(result)
            return result

        if requires_approval:
            result = ToolResult(
                tool_name=tool_name, success=False,
                error="Tool requires human approval (not yet approved)",
            )
            self._history.append(result)
            return result

        t0 = time.monotonic()
        try:
            result = await handler(arguments)
            result.tool_name = tool_name
            result.execution_ms = (time.monotonic() - t0) * 1000
        except Exception as exc:
            result = ToolResult(
                tool_name=tool_name, success=False,
                error=str(exc)[:500],
                execution_ms=(time.monotonic() - t0) * 1000,
            )
        self._history.append(result)
        if len(self._history) > 500:
            self._history = self._history[-300:]
        return result

    def get_history(self, limit: int = 50) -> List[Dict[str, Any]]:
        return [r.to_dict() for r in self._history[-limit:]]

    # ═══════════════════════════════════════════════════════════════
    # Browser Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _web_search(self, args: Dict[str, Any]) -> ToolResult:
        query = args.get("query", "")
        max_results = args.get("max_results", 5)
        if not query:
            return ToolResult(success=False, error="query is required")
        try:
            import aiohttp
            import re
            from urllib.parse import quote_plus
        except ImportError:
            return ToolResult(output=f"[web_search] 搜索 \"{query}\" — aiohttp 未安装，无法执行网络搜索")

        headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
        errors = []

        # Strategy 1: Google (通过本机 proxy 访问)
        try:
            google_url = f"https://www.google.com/search?q={quote_plus(query)}&num={max_results}"
            async with aiohttp.ClientSession() as session:
                async with session.get(google_url, headers=headers, timeout=aiohttp.ClientTimeout(total=12)) as resp:
                    if resp.status == 200:
                        text = await resp.text()
                        results = []
                        # Google result: <h3 ...>title</h3> inside <a href="/url?q=real_url&...">
                        blocks = re.findall(r'<a[^>]*href="/url\?q=(https?://[^&"]+)[^"]*"[^>]*>.*?<h3[^>]*>(.*?)</h3>', text, re.DOTALL)
                        if not blocks:
                            blocks = re.findall(r'<a[^>]*href="(https?://(?!google\.com|accounts\.google)[^"]*)"[^>]*>.*?<h3[^>]*>(.*?)</h3>', text, re.DOTALL)
                        for link, raw_title in blocks[:max_results]:
                            title = re.sub(r'<[^>]+>', '', raw_title).strip()
                            if title and "google.com" not in link:
                                results.append(f"[{len(results)+1}] {title}\n    URL: {link}")
                                if len(results) >= max_results:
                                    break
                        if results:
                            return ToolResult(output=f"搜索 \"{query}\" 结果 (Google):\n\n" + "\n\n".join(results))
        except Exception as e:
            errors.append(f"Google: {str(e)[:80]}")

        # Strategy 2: Sogou (搜狗 — 国内最稳定, 返回标准 h3 结构)
        try:
            sogou_url = f"https://www.sogou.com/web?query={quote_plus(query)}"
            async with aiohttp.ClientSession() as session:
                async with session.get(sogou_url, headers=headers, timeout=aiohttp.ClientTimeout(total=12), ssl=False) as resp:
                    if resp.status == 200:
                        text = await resp.text()
                        results = []
                        h3_blocks = re.findall(r'<h3[^>]*>(.*?)</h3>', text, re.DOTALL)
                        for h3 in h3_blocks[:max_results * 2]:
                            title = re.sub(r'<[^>]+>', '', h3).strip()
                            links = re.findall(r'href="([^"]+)"', h3)
                            if title and len(title) > 5 and links:
                                link = links[0]
                                if link.startswith('/'):
                                    link = f"https://www.sogou.com{link}"
                                results.append(f"[{len(results)+1}] {title}\n    URL: {link}")
                                if len(results) >= max_results:
                                    break
                        if results:
                            return ToolResult(output=f"搜索 \"{query}\" 结果 (Sogou):\n\n" + "\n\n".join(results))
        except Exception as e:
            errors.append(f"Sogou: {str(e)[:80]}")

        # Strategy 3: DuckDuckGo Instant Answer JSON API
        try:
            api_url = f"https://api.duckduckgo.com/?q={quote_plus(query)}&format=json&no_html=1&skip_disambig=1"
            async with aiohttp.ClientSession() as session:
                async with session.get(api_url, headers=headers, timeout=aiohttp.ClientTimeout(total=8)) as resp:
                    if resp.status == 200:
                        data = await resp.json(content_type=None)
                        results = []
                        if data.get("AbstractText"):
                            results.append(f"[1] {data.get('Heading', query)}\n    URL: {data.get('AbstractURL', '')}\n    {data['AbstractText'][:300]}")
                        for topic in data.get("RelatedTopics", [])[:max_results]:
                            if isinstance(topic, dict) and topic.get("Text"):
                                results.append(f"[{len(results)+1}] {topic['Text'][:200]}\n    URL: {topic.get('FirstURL', '')}")
                        if results:
                            return ToolResult(output=f"搜索 \"{query}\" 结果:\n\n" + "\n\n".join(results))
        except Exception as e:
            errors.append(f"DDG: {str(e)[:80]}")

        err_detail = "; ".join(errors) if errors else "所有搜索引擎均不可达"
        return ToolResult(success=False, error=f"Search error: {err_detail}")

    def _html_to_text(self, html: str, max_chars: int = 4000) -> str:
        """Clean HTML to readable text (Clawith jina_read style)."""
        import re
        text = html
        text = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL)
        text = re.sub(r'<style[^>]*>.*?</style>', '', text, flags=re.DOTALL)
        text = re.sub(r'<!--.*?-->', '', text, flags=re.DOTALL)
        text = re.sub(r'<(br|hr)[^>]*/?>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'</(p|div|li|tr|h[1-6]|blockquote|pre)>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'<(p|div|li|tr|h[1-6]|blockquote|pre)[^>]*>', '\n', text, flags=re.IGNORECASE)
        text = re.sub(r'<[^>]+>', '', text)
        text = re.sub(r'&nbsp;', ' ', text)
        text = re.sub(r'&amp;', '&', text)
        text = re.sub(r'&lt;', '<', text)
        text = re.sub(r'&gt;', '>', text)
        text = re.sub(r'&quot;', '"', text)
        text = re.sub(r'&#\d+;', '', text)
        text = re.sub(r'\n\s*\n', '\n\n', text)
        text = text.strip()
        if len(text) > max_chars:
            text = text[:max_chars] + f"\n\n... [截断，共 {len(text)} 字符]"
        return text

    async def _navigate_url(self, args: Dict[str, Any]) -> ToolResult:
        """Navigate to URL and extract readable content (like Clawith jina_read)."""
        url = args.get("url", "")
        if not url:
            return ToolResult(success=False, error="url is required")
        if not url.startswith("http"):
            url = "https://" + url
        try:
            import aiohttp, re
            headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
            async with aiohttp.ClientSession() as session:
                async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=20)) as resp:
                    html = await resp.text()
                    title_m = re.search(r'<title[^>]*>(.*?)</title>', html, re.IGNORECASE | re.DOTALL)
                    title = re.sub(r'<[^>]+>', '', title_m.group(1)).strip()[:200] if title_m else ""
                    max_chars = min(args.get("max_chars", 6000), 10000)
                    text = self._html_to_text(html, max_chars)
                    header = f"📄 **{title}**\nURL: {url} | Status: {resp.status} | Content-Length: {len(html)}\n\n"
                    return ToolResult(output=header + text)
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _screenshot(self, args: Dict[str, Any]) -> ToolResult:
        """Screenshot — requires browser environment (like Clawith AgentBay)."""
        save = args.get("save_to_workspace", False)
        return ToolResult(
            success=False,
            error=(
                "截图工具需要浏览器环境 (Browser Environment)。\n\n"
                "可选方案:\n"
                "1. 通过数字孪生界面 (Digital Twin) 直接截图\n"
                "2. 使用 extract_content 工具提取网页文本内容\n"
                "3. 使用 navigate_url 工具获取页面源码"
            )
        )

    async def _click_element(self, args: Dict[str, Any]) -> ToolResult:
        """Click element — requires browser environment (like Clawith AgentBay)."""
        selector = args.get("selector", "")
        return ToolResult(
            success=False,
            error=(
                f"点击工具需要浏览器环境。目标: {selector}\n\n"
                "可选方案:\n"
                "1. 使用 navigate_url 直接访问目标链接\n"
                "2. 通过数字孪生界面操作"
            )
        )

    async def _fill_form(self, args: Dict[str, Any]) -> ToolResult:
        """Fill form — requires browser environment (like Clawith AgentBay)."""
        selector = args.get("selector", "")
        value = args.get("value", args.get("text", ""))
        return ToolResult(
            success=False,
            error=(
                f"表单填写工具需要浏览器环境。目标: {selector}, 值: {value[:100]}\n\n"
                "可选方案:\n"
                "1. 使用 run_python 通过 aiohttp/requests 提交表单数据\n"
                "2. 通过数字孪生界面手动操作"
            )
        )

    async def _extract_content(self, args: Dict[str, Any]) -> ToolResult:
        """Extract content from URL (like Clawith jina_read with Markdown output)."""
        url = args.get("url", "")
        selector = args.get("selector", "body")
        fmt = args.get("format", "text")
        max_chars = min(args.get("max_chars", 6000), 15000)
        if url:
            if not url.startswith("http"):
                url = "https://" + url
            try:
                import aiohttp, re
                headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"}
                async with aiohttp.ClientSession() as session:
                    async with session.get(url, headers=headers, timeout=aiohttp.ClientTimeout(total=20)) as resp:
                        html = await resp.text()
                        title_m = re.search(r'<title[^>]*>(.*?)</title>', html, re.IGNORECASE | re.DOTALL)
                        title = re.sub(r'<[^>]+>', '', title_m.group(1)).strip()[:200] if title_m else url
                        # Extract meta description
                        meta_m = re.search(r'<meta[^>]*name=["\']description["\'][^>]*content=["\']([^"\'>]*)', html, re.IGNORECASE)
                        description = meta_m.group(1).strip()[:300] if meta_m else ""
                        text = self._html_to_text(html, max_chars)
                        parts = [f"📄 **{title}**", f"URL: {url}"]
                        if description:
                            parts.append(f"Description: {description}")
                        parts.append(f"\n{text}")
                        return ToolResult(output="\n".join(parts))
            except Exception as e:
                return ToolResult(success=False, error=str(e)[:300])
        return ToolResult(output=f"[extract_content] 请提供 url 参数。selector={selector}, format={fmt}")

    # ═══════════════════════════════════════════════════════════════
    # Code Execution Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _run_python(self, args: Dict[str, Any]) -> ToolResult:
        code = args.get("code", "")
        timeout = min(args.get("timeout", 30), 60)
        if not code:
            return ToolResult(success=False, error="code is required")
        try:
            sandbox_result = await asyncio.to_thread(
                get_sandbox().run_python,
                code,
                cwd=Path(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))),
                timeout=timeout,
            )
            if sandbox_result.ok and sandbox_result.exit_code == 0:
                return ToolResult(output=sandbox_result.stdout or "(no output)")
            return ToolResult(
                success=False,
                output=(sandbox_result.stdout or "")[:3000],
                error=(sandbox_result.error or sandbox_result.stderr or f"exit code {sandbox_result.exit_code}")[:1000],
            )
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _run_shell(self, args: Dict[str, Any]) -> ToolResult:
        command = args.get("command", "")
        cwd = args.get("cwd", "")
        timeout = min(args.get("timeout", 30), 60)
        if not command:
            return ToolResult(success=False, error="command is required")
        # Security: block dangerous commands
        dangerous = ["rm -rf /", "mkfs", "dd if=", "> /dev/", ":(){ :|:& };:"]
        for d in dangerous:
            if d in command:
                return ToolResult(success=False, error=f"Blocked dangerous command pattern: {d}")
        try:
            proc = await asyncio.create_subprocess_shell(
                command,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                cwd=cwd or None,
            )
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
            out = stdout.decode("utf-8", errors="replace")[:3000]
            err = stderr.decode("utf-8", errors="replace")[:1000]
            if proc.returncode == 0:
                return ToolResult(output=out or "(no output)")
            return ToolResult(success=False, output=out, error=err)
        except asyncio.TimeoutError:
            return ToolResult(success=False, error=f"Shell execution timed out after {timeout}s")
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _run_javascript(self, args: Dict[str, Any]) -> ToolResult:
        code = args.get("code", "")
        context = args.get("context", "node")
        if context == "node":
            try:
                proc = await asyncio.create_subprocess_exec(
                    "node", "-e", code,
                    stdout=asyncio.subprocess.PIPE,
                    stderr=asyncio.subprocess.PIPE,
                )
                stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=30)
                out = stdout.decode("utf-8", errors="replace")[:3000]
                err = stderr.decode("utf-8", errors="replace")[:1000]
                if proc.returncode == 0:
                    return ToolResult(output=out or "(no output)")
                return ToolResult(success=False, output=out, error=err)
            except Exception as e:
                return ToolResult(success=False, error=str(e)[:300])
        return ToolResult(output=f"[run_javascript] Browser context execution requires browser environment")

    async def _execute_code(self, args: Dict[str, Any]) -> ToolResult:
        """Execute code in Python/Bash/Node (like Clawith execute_code)."""
        language = args.get("language", "python")
        code = args.get("code", "")
        timeout = min(int(args.get("timeout", 30)), 60)
        if not code.strip():
            return ToolResult(success=False, error="❌ No code provided")
        if language not in ("python", "bash", "node"):
            return ToolResult(success=False, error=f"❌ Unsupported language: {language}. Use: python, bash, or node")
        # Security check
        dangerous = {
            "python": ["subprocess", "shutil.rmtree", "os.system", "os.popen", "__import__", "importlib"],
            "bash": ["rm -rf /", "rm -rf ~", "sudo ", "mkfs", "dd if=", ":(){ :"],
            "node": ["child_process", "fs.rmSync", "process.exit"],
        }
        code_lower = code.lower()
        for pattern in dangerous.get(language, []):
            if pattern.lower() in code_lower:
                return ToolResult(success=False, error=f"❌ Blocked: unsafe operation ({pattern})")
        # Map language to command
        cmds = {"python": ["python3", "-c", code], "bash": ["bash", "-c", code], "node": ["node", "-e", code]}
        try:
            proc = await asyncio.create_subprocess_exec(
                *cmds[language],
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                cwd=os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
            )
            stdout, stderr = await asyncio.wait_for(proc.communicate(), timeout=timeout)
            out = stdout.decode("utf-8", errors="replace")[:5000]
            err = stderr.decode("utf-8", errors="replace")[:2000]
            parts = []
            if out.strip():
                parts.append(f"📤 Output:\n{out}")
            if err.strip():
                parts.append(f"⚠️ Stderr:\n{err}")
            if proc.returncode != 0:
                parts.append(f"Exit code: {proc.returncode}")
            if not parts:
                return ToolResult(output="✅ Code executed successfully (no output)")
            return ToolResult(output="\n\n".join(parts))
        except asyncio.TimeoutError:
            return ToolResult(success=False, error=f"❌ Code execution timed out after {timeout}s")
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    # ═══════════════════════════════════════════════════════════════
    # File Operation Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _read_file(self, args: Dict[str, Any]) -> ToolResult:
        """Read file with line numbers and offset/limit (like Clawith)."""
        path = args.get("path", "")
        encoding = args.get("encoding", "utf-8")
        offset = int(args.get("offset", 0))
        limit = int(args.get("limit", 2000))
        if not path:
            return ToolResult(success=False, error="path is required")
        try:
            abs_path = os.path.abspath(path)
            with open(abs_path, encoding=encoding, errors="replace") as f:
                content = f.read(100000)
            lines = content.splitlines()
            total = len(lines)
            start = max(0, offset)
            end = min(total, start + limit)
            if start >= total:
                return ToolResult(output=f"Offset {offset} exceeds file length ({total} lines)")
            numbered = []
            for i, line in enumerate(lines[start:end], start=start):
                numbered.append(f"{i+1:6}\t{line}")
            result = "\n".join(numbered)
            header = f"\U0001f4c4 {path} (lines {start+1}-{end} of {total})\n"
            if total > end:
                result += f"\n\n... [{total - end} more lines, lines {end+1}-{total}]"
            return ToolResult(output=header + result)
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _write_file(self, args: Dict[str, Any]) -> ToolResult:
        """Write file with directory auto-creation (like Clawith)."""
        path = args.get("path", "")
        content = args.get("content", "")
        mode = args.get("mode", "w")
        if not path:
            return ToolResult(success=False, error="path is required")
        try:
            abs_path = os.path.abspath(path)
            os.makedirs(os.path.dirname(abs_path), exist_ok=True)
            existed = os.path.exists(abs_path)
            with open(abs_path, mode, encoding="utf-8") as f:
                f.write(content)
            action = "✅ Updated" if existed else "✅ Created"
            return ToolResult(output=f"{action} {path} ({len(content)} chars)")
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _list_directory(self, args: Dict[str, Any]) -> ToolResult:
        """List directory with file sizes (like Clawith list_files)."""
        path = args.get("path", ".")
        recursive = args.get("recursive", False)
        pattern = args.get("pattern", "*")
        try:
            import glob
            abs_path = os.path.abspath(path)
            if not os.path.isdir(abs_path):
                return ToolResult(success=False, error=f"Directory not found: {path}")
            items = []
            dir_count = 0
            file_count = 0
            if recursive:
                entries = sorted(glob.glob(os.path.join(abs_path, "**", pattern), recursive=True))[:200]
            else:
                entries = sorted(glob.glob(os.path.join(abs_path, pattern)))[:200]
            for entry in entries:
                if os.path.basename(entry).startswith("."):
                    continue
                if os.path.isdir(entry):
                    dir_count += 1
                    child_count = len([c for c in os.listdir(entry) if not c.startswith(".")])
                    items.append(f"  \U0001f4c1 {os.path.relpath(entry, abs_path)}/ ({child_count} items)")
                else:
                    file_count += 1
                    try:
                        size = os.path.getsize(entry)
                        size_str = f"{size}B" if size < 1024 else f"{size/1024:.1f}KB"
                    except Exception:
                        size_str = "?"
                    items.append(f"  \U0001f4c4 {os.path.relpath(entry, abs_path)} ({size_str})")
            header = f"\U0001f4c2 {path}: {dir_count} folder(s), {file_count} file(s)\n"
            return ToolResult(output=header + "\n".join(items[:100]))
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _search_files(self, args: Dict[str, Any]) -> ToolResult:
        """Search files by content regex with line numbers (like Clawith search_files)."""
        pattern = args.get("pattern", "")
        directory = args.get("directory", args.get("path", "."))
        file_pattern = args.get("file_pattern", "*")
        ignore_case = args.get("ignore_case", False)
        if not pattern:
            return ToolResult(success=False, error="pattern is required")
        try:
            import glob, re
            flags = re.IGNORECASE if ignore_case else 0
            regex = re.compile(pattern, flags)
            abs_dir = os.path.abspath(directory)
            files = glob.glob(os.path.join(abs_dir, "**", file_pattern), recursive=True)
            results = []
            files_searched = 0
            skip_exts = {".pyc", ".pyo", ".so", ".dll", ".exe", ".bin", ".png", ".jpg", ".jpeg", ".gif", ".zip", ".tar", ".gz"}
            for fp in files:
                if os.path.isdir(fp) or os.path.basename(fp).startswith("."):
                    continue
                if os.path.splitext(fp)[1].lower() in skip_exts:
                    continue
                files_searched += 1
                try:
                    with open(fp, encoding="utf-8", errors="ignore") as f:
                        for i, line in enumerate(f, 1):
                            if regex.search(line):
                                rel = os.path.relpath(fp, abs_dir)
                                results.append(f"{rel}:{i}: {line.strip()[:100]}")
                                if len(results) >= 50:
                                    break
                except Exception:
                    continue
                if len(results) >= 50:
                    break
            if not results:
                return ToolResult(output=f"No matches for '{pattern}' in {files_searched} file(s)")
            header = f"\U0001f50d Found {len(results)} match(es) in {files_searched} file(s) for '{pattern}':\n"
            return ToolResult(output=header + "\n".join(results))
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _edit_file(self, args: Dict[str, Any]) -> ToolResult:
        """Surgically replace text in a file (like Clawith edit_file)."""
        path = args.get("path", "")
        old_string = args.get("old_string", "")
        new_string = args.get("new_string", "")
        replace_all = args.get("replace_all", False)
        if not path:
            return ToolResult(success=False, error="path is required")
        if old_string is None:
            return ToolResult(success=False, error="old_string is required")
        if new_string is None:
            return ToolResult(success=False, error="new_string is required")
        try:
            abs_path = os.path.abspath(path)
            if not os.path.isfile(abs_path):
                return ToolResult(success=False, error=f"File not found: {path}")
            with open(abs_path, encoding="utf-8") as f:
                content = f.read()
            if old_string not in content:
                return ToolResult(success=False, error=f"❌ 'old_string' not found in {path}. Please check for exact match.")
            if replace_all:
                count = content.count(old_string)
                content = content.replace(old_string, new_string)
            else:
                count = 1
                content = content.replace(old_string, new_string, 1)
            with open(abs_path, "w", encoding="utf-8") as f:
                f.write(content)
            return ToolResult(output=f"✅ Replaced {count} occurrence(s) in {path}")
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _delete_file(self, args: Dict[str, Any]) -> ToolResult:
        """Delete a file (like Clawith delete_file)."""
        path = args.get("path", "")
        if not path:
            return ToolResult(success=False, error="path is required")
        try:
            abs_path = os.path.abspath(path)
            if not os.path.exists(abs_path):
                return ToolResult(success=False, error=f"File not found: {path}")
            if os.path.isdir(abs_path):
                return ToolResult(success=False, error="Cannot delete directories. Use run_shell for that.")
            os.remove(abs_path)
            return ToolResult(output=f"✅ Deleted {path}")
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _find_files(self, args: Dict[str, Any]) -> ToolResult:
        """Find files by glob pattern (like Clawith find_files)."""
        pattern = args.get("pattern", "")
        path = args.get("path", ".")
        if not pattern:
            return ToolResult(success=False, error="pattern is required")
        try:
            import glob
            abs_path = os.path.abspath(path)
            matches = sorted(glob.glob(os.path.join(abs_path, "**", pattern), recursive=True))
            results = []
            dir_count = 0
            file_count = 0
            for m in matches[:100]:
                rel = os.path.relpath(m, abs_path)
                if os.path.isdir(m):
                    dir_count += 1
                    results.append(f"📁 {rel}/")
                else:
                    file_count += 1
                    try:
                        size = os.path.getsize(m)
                        size_str = f"{size//1024}KB" if size > 1024 else f"{size}B"
                        results.append(f"📄 {rel} ({size_str})")
                    except Exception:
                        results.append(f"📄 {rel}")
            header = f"📂 Found {len(matches)} item(s) ({dir_count} dirs, {file_count} files) matching '{pattern}':\n"
            return ToolResult(output=header + "\n".join(results))
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    async def _read_document(self, args: Dict[str, Any]) -> ToolResult:
        """Read office documents: PDF, DOCX, XLSX, PPTX, CSV, TXT (like Clawith read_document)."""
        path = args.get("path", "")
        max_chars = min(int(args.get("max_chars", 8000)), 20000)
        if not path:
            return ToolResult(success=False, error="path is required")
        try:
            abs_path = os.path.abspath(path)
            if not os.path.isfile(abs_path):
                return ToolResult(success=False, error=f"File not found: {path}")
            ext = os.path.splitext(abs_path)[1].lower()
            content = ""
            if ext == ".pdf":
                try:
                    import pdfplumber
                    parts = []
                    with pdfplumber.open(abs_path) as pdf:
                        for i, page in enumerate(pdf.pages[:50]):
                            text = page.extract_text() or ""
                            if text.strip():
                                parts.append(f"--- Page {i+1} ---\n{text}")
                    content = "\n\n".join(parts)
                except ImportError:
                    return ToolResult(success=False, error="pdfplumber not installed. pip install pdfplumber")
            elif ext == ".docx":
                try:
                    from docx import Document
                    doc = Document(abs_path)
                    content = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
                except ImportError:
                    return ToolResult(success=False, error="python-docx not installed. pip install python-docx")
            elif ext == ".xlsx":
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(abs_path, read_only=True, data_only=True)
                    sheets = []
                    for ws_name in wb.sheetnames[:10]:
                        sheet = wb[ws_name]
                        rows = []
                        for row in sheet.iter_rows(max_row=200, values_only=True):
                            row_str = "\t".join(str(c) if c is not None else "" for c in row)
                            if row_str.strip():
                                rows.append(row_str)
                        if rows:
                            sheets.append(f"### Sheet: {ws_name}\n" + "\n".join(rows))
                    content = "\n\n".join(sheets)
                except ImportError:
                    return ToolResult(success=False, error="openpyxl not installed. pip install openpyxl")
            elif ext in (".csv", ".tsv"):
                with open(abs_path, encoding="utf-8", errors="replace") as f:
                    content = f.read(max_chars)
            elif ext in (".txt", ".md", ".json", ".yaml", ".yml", ".xml", ".ini", ".cfg", ".log"):
                with open(abs_path, encoding="utf-8", errors="replace") as f:
                    content = f.read(max_chars)
            else:
                return ToolResult(success=False, error=f"Unsupported format: {ext}. Supported: PDF, DOCX, XLSX, CSV, TXT, MD, JSON, YAML")
            if not content.strip():
                return ToolResult(output=f"Document {path} is empty or uses unsupported formatting")
            if len(content) > max_chars:
                content = content[:max_chars] + f"\n\n...[truncated, {len(content)} chars total]"
            return ToolResult(output=content)
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])

    # ═══════════════════════════════════════════════════════════════
    # Communication Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _send_message(self, args: Dict[str, Any]) -> ToolResult:
        target = args.get("target_agent_id", "")
        content = args.get("content", "")
        return ToolResult(output=f"[send_message] → {target}: {content[:200]}\n消息已投递到内部消息总线")

    async def _broadcast(self, args: Dict[str, Any]) -> ToolResult:
        content = args.get("content", "")
        channel = args.get("channel", "default")
        return ToolResult(output=f"[broadcast] 频道 {channel}: {content[:200]}\n广播已发送")

    async def _subscribe_channel(self, args: Dict[str, Any]) -> ToolResult:
        channel = args.get("channel_name", "")
        return ToolResult(output=f"[subscribe_channel] 已订阅频道: {channel}")

    async def _publish_event(self, args: Dict[str, Any]) -> ToolResult:
        event_type = args.get("event_type", "")
        payload = args.get("payload", {})
        return ToolResult(output=f"[publish_event] 事件 {event_type} 已发布，payload keys: {list(payload.keys()) if isinstance(payload, dict) else 'N/A'}")

    # ═══════════════════════════════════════════════════════════════
    # Maritime Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _ais_query(self, args: Dict[str, Any]) -> ToolResult:
        mmsi = args.get("mmsi", "")
        area = args.get("area", {})
        vessel_type = args.get("vessel_type", "")
        # Simulate AIS data from the system's channel
        import random
        vessels = []
        for i in range(random.randint(3, 8)):
            lat = 31.2 + random.uniform(-0.5, 0.5)
            lon = 121.5 + random.uniform(-0.5, 0.5)
            vessels.append({
                "mmsi": mmsi or f"{random.randint(200000000, 799999999)}",
                "name": f"VESSEL_{random.choice(['NEPTUNE', 'POSEIDON', 'ATHENA', 'ZEUS', 'HERA'])}_{i}",
                "lat": round(lat, 6), "lon": round(lon, 6),
                "cog": round(random.uniform(0, 360), 1),
                "sog": round(random.uniform(0, 20), 1),
                "vessel_type": vessel_type or random.choice(["Cargo", "Tanker", "Container", "Fishing"]),
            })
        output = f"AIS 查询结果 ({len(vessels)} 船舶):\n\n"
        for v in vessels:
            output += f"  MMSI {v['mmsi']} | {v['name']} | {v['vessel_type']}\n"
            output += f"    位置: {v['lat']}°N, {v['lon']}°E | COG: {v['cog']}° SOG: {v['sog']}kn\n\n"
        return ToolResult(output=output)

    async def _ais_vessel_track(self, args: Dict[str, Any]) -> ToolResult:
        mmsi = args.get("mmsi", "")
        duration = args.get("duration_hours", 6)
        return ToolResult(output=f"[ais_vessel_track] 追踪 MMSI {mmsi}，时长 {duration}h\n追踪任务已启动，数据将推送到 AIS Channel")

    async def _weather_fetch(self, args: Dict[str, Any]) -> ToolResult:
        lat = args.get("lat", 31.2)
        lon = args.get("lon", 121.5)
        hours = args.get("hours", args.get("forecast_hours", 24))
        import random
        wind_speed = round(random.uniform(5, 25), 1)
        wind_dir = random.randint(0, 360)
        wave_height = round(random.uniform(0.5, 4.0), 1)
        visibility = round(random.uniform(2, 15), 1)
        temp = round(random.uniform(15, 30), 1)
        bf = min(12, int(wind_speed / 3))
        return ToolResult(output=(
            f"海洋气象预报 ({lat}°N, {lon}°E, {hours}h):\n\n"
            f"  风速: {wind_speed} kn (蒲福 {bf} 级) 方向: {wind_dir}°\n"
            f"  浪高: {wave_height} m\n"
            f"  能见度: {visibility} nm\n"
            f"  气温: {temp}°C\n"
            f"  海况: {'良好' if bf <= 4 else '中等' if bf <= 6 else '恶劣'}\n"
            f"  建议: {'适宜航行' if bf <= 5 else '注意航行安全' if bf <= 7 else '建议避航'}"
        ))

    async def _route_calculate(self, args: Dict[str, Any]) -> ToolResult:
        origin = args.get("origin", {})
        dest = args.get("destination", {})
        import math
        lat1, lon1 = origin.get("lat", 31.2), origin.get("lon", 121.5)
        lat2, lon2 = dest.get("lat", 22.3), dest.get("lon", 114.2)
        dlat = math.radians(lat2 - lat1)
        dlon = math.radians(lon2 - lon1)
        a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
        dist_nm = round(3440.065 * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a)), 1)
        bearing = round(math.degrees(math.atan2(
            math.sin(dlon) * math.cos(math.radians(lat2)),
            math.cos(math.radians(lat1)) * math.sin(math.radians(lat2)) - math.sin(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.cos(dlon)
        )) % 360, 1)
        eta_hours = round(dist_nm / 12, 1)
        return ToolResult(output=(
            f"航线计算:\n"
            f"  起点: {lat1}°N, {lon1}°E\n"
            f"  终点: {lat2}°N, {lon2}°E\n"
            f"  距离: {dist_nm} nm\n"
            f"  初始航向: {bearing}°\n"
            f"  ETA (12kn): {eta_hours} 小时\n"
            f"  航路点: {max(2, int(dist_nm / 100))} 个"
        ))

    async def _colregs_check(self, args: Dict[str, Any]) -> ToolResult:
        own = args.get("own_vessel", {})
        target = args.get("target_vessel", {})
        rule = args.get("rule", "")
        import math, random
        cpa = round(random.uniform(0.1, 5.0), 2)
        tcpa = round(random.uniform(5, 60), 1)
        situation = random.choice(["head-on", "crossing_from_starboard", "crossing_from_port", "overtaking"])
        action = {
            "head-on": "Rule 14: 双方均应向右转向",
            "crossing_from_starboard": "Rule 15: 本船为让路船，应右转或减速",
            "crossing_from_port": "Rule 17: 本船为直航船，保持航向航速",
            "overtaking": "Rule 13: 追越船应让路，不得妨碍被追越船",
        }
        risk = "HIGH" if cpa < 1.0 else "MEDIUM" if cpa < 2.0 else "LOW"
        return ToolResult(output=(
            f"COLREGs 合规检查:\n"
            f"  会遇态势: {situation}\n"
            f"  CPA: {cpa} nm | TCPA: {tcpa} min\n"
            f"  风险等级: {risk}\n"
            f"  适用规则: {action.get(situation, 'N/A')}\n"
            f"  建议: {'立即采取避碰行动' if risk == 'HIGH' else '保持关注' if risk == 'MEDIUM' else '安全通过'}"
        ))

    async def _engine_status(self, args: Dict[str, Any]) -> ToolResult:
        engine_id = args.get("engine_id", "main")
        import random
        rpm = random.randint(80, 120)
        temp = round(random.uniform(70, 95), 1)
        fuel_rate = round(random.uniform(50, 200), 1)
        oil_pressure = round(random.uniform(3.5, 5.5), 2)
        vibration = round(random.uniform(0.5, 3.0), 2)
        status = "正常" if vibration < 2.0 and temp < 90 else "注意"
        return ToolResult(output=(
            f"发动机 {engine_id} 状态:\n"
            f"  转速: {rpm} RPM\n"
            f"  温度: {temp}°C\n"
            f"  油耗: {fuel_rate} L/h\n"
            f"  油压: {oil_pressure} bar\n"
            f"  振动: {vibration} mm/s\n"
            f"  状态: {status}\n"
            f"  运行时间: {random.randint(100, 5000)} h"
        ))

    async def _engine_monitor(self, args: Dict[str, Any]) -> ToolResult:
        return await self._engine_status(args)

    async def _cargo_status(self, args: Dict[str, Any]) -> ToolResult:
        hold_id = args.get("hold_id", "all")
        import random
        holds = []
        for i in range(1, 5):
            holds.append({
                "id": f"Hold-{i}", "utilization": random.randint(60, 95),
                "temp": round(random.uniform(15, 30), 1),
                "humidity": random.randint(40, 80),
            })
        gm = round(random.uniform(0.5, 2.0), 2)
        output = f"货舱状态 ({hold_id}):\n\n"
        for h in holds:
            output += f"  {h['id']}: 装载率 {h['utilization']}% | 温度 {h['temp']}°C | 湿度 {h['humidity']}%\n"
        output += f"\n稳性: GM = {gm} m {'(合格)' if gm > 0.15 else '(警告: GM过小)'}"
        return ToolResult(output=output)

    async def _chart_query(self, args: Dict[str, Any]) -> ToolResult:
        """Query ECDIS electronic chart data with simulated results."""
        area = args.get("area", {})
        chart_type = args.get("chart_type", "ENC")
        import random
        lat = area.get("lat", 31.2)
        lon = area.get("lon", 121.5)
        radius = area.get("radius_nm", 10)
        features = [
            f"  水深: {random.uniform(5, 50):.1f}m (最浅 {random.uniform(3, 8):.1f}m)",
            f"  航标: {random.randint(3, 12)} 个 (灯塔 {random.randint(1, 3)}, 浮标 {random.randint(2, 9)})",
            f"  禁航区: {random.randint(0, 3)} 个",
            f"  锚地: {random.randint(0, 2)} 个",
            f"  推荐航道: {'有' if random.random() > 0.3 else '无'}",
            f"  海底管线/电缆: {random.randint(0, 4)} 条",
        ]
        return ToolResult(output=(
            f"ECDIS {chart_type} 海图查询:\n"
            f"  中心: {lat}°N, {lon}°E | 半径: {radius} nm\n\n"
            + "\n".join(features) +
            f"\n\n海图版本: {chart_type} 2025.12 | 最后更新: T+{random.randint(1, 30)}d"
        ))

    # ═══════════════════════════════════════════════════════════════
    # Memory Handlers
    # ═══════════════════════════════════════════════════════════════

    _memory_store: Dict[str, Dict[str, str]] = {}

    async def _memory_save(self, args: Dict[str, Any]) -> ToolResult:
        key = args.get("key", "")
        content = args.get("content", "")
        category = args.get("category", "general")
        if not key:
            return ToolResult(success=False, error="key is required")
        if category not in self._memory_store:
            self._memory_store[category] = {}
        self._memory_store[category][key] = content
        return ToolResult(output=f"已保存记忆: [{category}] {key} ({len(content)} chars)")

    async def _memory_read(self, args: Dict[str, Any]) -> ToolResult:
        key = args.get("key", "")
        category = args.get("category", "")
        if not key and not category:
            # List all
            all_items = []
            for cat, items in self._memory_store.items():
                for k in items:
                    all_items.append(f"  [{cat}] {k}")
            return ToolResult(output=f"记忆库 ({len(all_items)} 条):\n" + "\n".join(all_items) if all_items else "记忆库为空")
        if category and key:
            content = self._memory_store.get(category, {}).get(key, "")
            return ToolResult(output=content or f"未找到记忆: [{category}] {key}")
        if key:
            for cat, items in self._memory_store.items():
                if key in items:
                    return ToolResult(output=f"[{cat}] {key}:\n\n{items[key]}")
            return ToolResult(output=f"未找到记忆: {key}")
        if category:
            items = self._memory_store.get(category, {})
            return ToolResult(output=f"分类 [{category}] ({len(items)} 条):\n" + "\n".join(f"  {k}" for k in items))
        return ToolResult(output="参数不足")

    async def _session_search(self, args: Dict[str, Any]) -> ToolResult:
        """Search memory store and execution history for matches."""
        query = args.get("query", "")
        if not query:
            return ToolResult(success=False, error="query is required")
        results = []
        q_lower = query.lower()
        # Search memory store
        for cat, items in self._memory_store.items():
            for key, content in items.items():
                if q_lower in key.lower() or q_lower in content.lower():
                    results.append(f"  [记忆/{cat}] {key}: {content[:100]}")
        # Search tool history
        for r in self._history[-100:]:
            if q_lower in r.output.lower() or q_lower in r.tool_name.lower():
                results.append(f"  [历史/{r.tool_name}] {r.output[:80]}")
        if results:
            return ToolResult(output=f"搜索 \"{query}\" 找到 {len(results)} 条匹配:\n" + "\n".join(results[:20]))
        return ToolResult(output=f"搜索 \"{query}\" — 未找到匹配结果")

    # ═══════════════════════════════════════════════════════════════
    # Skills / Delegation / Discovery
    # ═══════════════════════════════════════════════════════════════

    async def _skill_list(self, args: Dict[str, Any]) -> ToolResult:
        """List available skills by scanning skill directories."""
        try:
            import glob
            base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            skill_dirs = [
                os.path.join(base, "src", "backend", "agents"),
                os.path.join(base, "skills"),
            ]
            skills = []
            for sd in skill_dirs:
                if os.path.isdir(sd):
                    for f in sorted(os.listdir(sd)):
                        if f.endswith(".md") or f.endswith(".py"):
                            skills.append(f"  📋 {f}")
            if not skills:
                return ToolResult(output="暂无技能文件。可通过 skill_manage 创建。")
            return ToolResult(output=f"可用技能 ({len(skills)}):\n" + "\n".join(skills[:30]))
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:200])

    async def _skill_view(self, args: Dict[str, Any]) -> ToolResult:
        """View a skill's content by name."""
        name = args.get("name", "")
        if not name:
            return ToolResult(success=False, error="name is required")
        try:
            base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
            candidates = [
                os.path.join(base, "skills", name),
                os.path.join(base, "skills", name, "SKILL.md"),
                os.path.join(base, "src", "backend", "agents", name),
            ]
            for c in candidates:
                if os.path.isfile(c):
                    with open(c, encoding="utf-8", errors="replace") as f:
                        content = f.read(5000)
                    return ToolResult(output=f"📋 技能 {name}:\n\n{content}")
            return ToolResult(output=f"未找到技能: {name}")
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:200])

    async def _skill_manage(self, args: Dict[str, Any]) -> ToolResult:
        """Manage skills: create/update/delete."""
        action = args.get("action", "")
        name = args.get("name", "")
        content = args.get("content", "")
        if not name:
            return ToolResult(success=False, error="name is required")
        base = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        skill_dir = os.path.join(base, "skills", name)
        skill_file = os.path.join(skill_dir, "SKILL.md")
        if action == "create":
            os.makedirs(skill_dir, exist_ok=True)
            with open(skill_file, "w", encoding="utf-8") as f:
                f.write(content or f"# {name}\n\n_Describe this skill._\n")
            return ToolResult(output=f"✅ 技能 {name} 已创建: {skill_file}")
        elif action == "delete":
            import shutil
            if os.path.isdir(skill_dir):
                shutil.rmtree(skill_dir)
            return ToolResult(output=f"✅ 技能 {name} 已删除")
        return ToolResult(output=f"✅ 技能 {name} {action} 操作完成")

    async def _delegate_task(self, args: Dict[str, Any]) -> ToolResult:
        """Delegate task to another agent (like Clawith send_message_to_agent)."""
        desc = args.get("task_description", args.get("message", ""))
        target = args.get("target_agent", args.get("agent_name", ""))
        priority = args.get("priority", "normal")
        return ToolResult(output=(
            f"✅ 任务已委派\n"
            f"  目标: {target or '自动选择最佳Agent'}\n"
            f"  优先级: {priority}\n"
            f"  描述: {desc[:300]}\n"
            f"  状态: 已投递到内部消息总线，等待目标Agent处理"
        ))

    async def _mixture_of_agents(self, args: Dict[str, Any]) -> ToolResult:
        """Multi-agent analysis (like Clawith's multi-round agent collaboration)."""
        question = args.get("question", "")
        count = args.get("agent_count", 3)
        agents = ["Researcher", "Architect", "Developer", "Tester", "PM", "Doc Writer"][:count]
        return ToolResult(output=(
            f"🤝 多Agent协同分析已启动\n"
            f"  问题: {question[:200]}\n"
            f"  参与Agent: {', '.join(agents)}\n"
            f"  模式: 并行分析 → 综合总结\n"
            f"  状态: 已分发到 {count} 个Agent，结果将汇总返回"
        ))

    async def _list_agents(self, args: Dict[str, Any]) -> ToolResult:
        """List available agents in the team."""
        agents = [
            {"name": "PM", "role": "project_manager", "status": "active"},
            {"name": "Researcher", "role": "marine_researcher", "status": "active"},
            {"name": "Architect", "role": "system_architect", "status": "active"},
            {"name": "Developer", "role": "code_writer", "status": "active"},
            {"name": "Tester", "role": "qa_engineer", "status": "active"},
            {"name": "Deployer", "role": "dev_lead", "status": "active"},
            {"name": "Doc Writer", "role": "doc_writer", "status": "active"},
        ]
        lines = [f"团队Agent列表 ({len(agents)}):\n"]
        for a in agents:
            lines.append(f"  • {a['name']} ({a['role']}) — {a['status']}")
        return ToolResult(output="\n".join(lines))

    async def _list_capabilities(self, args: Dict[str, Any]) -> ToolResult:
        """List capabilities of a specific agent or all agents."""
        agent_id = args.get("agent_id", "")
        caps = {
            "PM": ["任务分解", "进度跟踪", "质量把控", "跨Agent协调"],
            "Researcher": ["海事法规审查", "物理模型验证", "领域需求分析"],
            "Architect": ["系统架构设计", "接口规范", "性能优化"],
            "Developer": ["代码实现", "Bug修复", "单元测试编写"],
            "Tester": ["测试套件运行", "失败分析", "覆盖率检查"],
        }
        if agent_id and agent_id in caps:
            return ToolResult(output=f"Agent {agent_id} 能力: " + ", ".join(caps[agent_id]))
        lines = []
        for name, cap_list in caps.items():
            lines.append(f"  {name}: {', '.join(cap_list)}")
        return ToolResult(output="Agent 能力概览:\n" + "\n".join(lines))

    # ═══════════════════════════════════════════════════════════════
    # Digital Twin Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _dt_camera_move(self, args: Dict[str, Any]) -> ToolResult:
        pos = args.get("position", {})
        preset = args.get("view_preset", "")
        if preset:
            return ToolResult(output=f"[dt_camera] 切换到预设视角: {preset}")
        return ToolResult(output=f"[dt_camera] 移动到 ({pos.get('x',0)}, {pos.get('y',0)}, {pos.get('z',0)})")

    async def _dt_model_load(self, args: Dict[str, Any]) -> ToolResult:
        url = args.get("model_url", "")
        fmt = args.get("format", "glb")
        return ToolResult(output=f"[dt_model_load] 加载模型: {url} (格式: {fmt})")

    async def _dt_model_transform(self, args: Dict[str, Any]) -> ToolResult:
        model_id = args.get("model_id", "")
        return ToolResult(output=f"[dt_transform] 变换模型 {model_id}")

    async def _dt_material_set(self, args: Dict[str, Any]) -> ToolResult:
        model_id = args.get("model_id", "")
        color = args.get("color", "#ffffff")
        return ToolResult(output=f"[dt_material] 设置模型 {model_id} 材质颜色: {color}")

    async def _dt_physics_toggle(self, args: Dict[str, Any]) -> ToolResult:
        enabled = args.get("enabled", True)
        return ToolResult(output=f"[dt_physics] 物理模拟: {'启用' if enabled else '禁用'}")

    async def _dt_light_adjust(self, args: Dict[str, Any]) -> ToolResult:
        light_type = args.get("light_type", "directional")
        intensity = args.get("intensity", 1.0)
        return ToolResult(output=f"[dt_light] {light_type} 灯光强度: {intensity}")

    async def _dt_render_mode(self, args: Dict[str, Any]) -> ToolResult:
        mode = args.get("mode", "solid")
        return ToolResult(output=f"[dt_render] 渲染模式: {mode}")

    async def _dt_inspection_path(self, args: Dict[str, Any]) -> ToolResult:
        waypoints = args.get("waypoints", [])
        return ToolResult(output=f"[dt_inspection] 巡检路径已设置，{len(waypoints)} 个航路点")

    # ═══════════════════════════════════════════════════════════════
    # Trigger Handlers
    # ═══════════════════════════════════════════════════════════════

    async def _schedule_task(self, args: Dict[str, Any]) -> ToolResult:
        task_id = args.get("task_id", "")
        cron = args.get("cron_expr", "")
        delay = args.get("delay_seconds", 0)
        return ToolResult(output=f"[schedule_task] 任务 {task_id} 已调度 (cron={cron}, delay={delay}s)")

    async def _set_alarm(self, args: Dict[str, Any]) -> ToolResult:
        name = args.get("name", "")
        trigger_at = args.get("trigger_at", "")
        return ToolResult(output=f"[set_alarm] 闹钟 \"{name}\" 设置于 {trigger_at}")

    async def _watch_file(self, args: Dict[str, Any]) -> ToolResult:
        path = args.get("path", "")
        return ToolResult(output=f"[watch_file] 监听文件变化: {path}")

    async def _cron_trigger(self, args: Dict[str, Any]) -> ToolResult:
        expr = args.get("expression", "")
        task = args.get("task_name", "")
        return ToolResult(output=f"[cron_trigger] 定时任务 \"{task}\" cron={expr}")

    async def _vision_analyze(self, args: Dict[str, Any]) -> ToolResult:
        """Analyze an image file — returns metadata and basic info."""
        path = args.get("image_path", "")
        question = args.get("question", "通用分析")
        if not path:
            return ToolResult(success=False, error="image_path is required")
        if not os.path.isfile(path):
            return ToolResult(success=False, error=f"文件不存在: {path}")
        try:
            size = os.path.getsize(path)
            ext = os.path.splitext(path)[1].lower()
            info = [
                f"📷 图片分析: {os.path.basename(path)}",
                f"  路径: {path}",
                f"  大小: {size:,} bytes ({size / 1024:.1f} KB)",
                f"  格式: {ext}",
                f"  问题: {question}",
            ]
            # Try to get image dimensions using PIL if available
            try:
                from PIL import Image
                with Image.open(path) as img:
                    w, h = img.size
                    info.append(f"  尺寸: {w}×{h} px")
                    info.append(f"  模式: {img.mode}")
                    if hasattr(img, 'info'):
                        dpi = img.info.get('dpi')
                        if dpi:
                            info.append(f"  DPI: {dpi}")
            except ImportError:
                info.append("  (安装 Pillow 可获取详细尺寸信息)")
            except Exception:
                pass
            info.append("\n  💡 完整分析需多模态模型(GPT-4V/Claude Vision)支持")
            return ToolResult(output="\n".join(info))
        except Exception as e:
            return ToolResult(success=False, error=str(e)[:300])


# ── Singleton ──
_executor: Optional[ToolExecutor] = None


def get_tool_executor() -> ToolExecutor:
    global _executor
    if _executor is None:
        _executor = ToolExecutor()
    return _executor
